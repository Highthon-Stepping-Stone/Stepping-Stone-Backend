from requests import get
from bs4 import BeautifulSoup
import re

from openpyxl import load_workbook

from app.models.album import FreeAlbumModel, ScheduledAlbumModel
from app.models.school import SchoolModel

_url = 'http://{0}/sts_sci_sf00_001.do?schulCode={1}&schulCrseScCode=4&schulKndScScore=04&ay={2}&mm={3:0>2}'

WEB_URLS = {
    '서울특별시': 'stu.sen.go.kr',
    '부산광역시': 'stu.pen.go.kr',
    '대구광역시': 'stu.dge.go.kr',
    '인천광역시': 'stu.ice.go.kr',
    '광주광역시': 'stu.gen.go.kr',
    '대전광역시': 'stu.dje.go.kr',
    '울산광역시': 'stu.use.go.kr',
    '세종특별자치시': 'stu.sje.go.kr',
    '경기도': 'stu.cbe.go.kr',
    '강원도': 'stu.kwe.go.kr',
    '충청북도': 'stu.cbe.go.kr',
    '충청남도': 'stu.cne.go.kr',
    '전라북도': 'stu.jbe.go.kr',
    '전라남도': 'stu.jne.go.kr',
    '경상북도': 'stu.gbe.go.kr',
    '경상남도': 'stu.gne.go.kr',
    '제주특별자치도': 'stu.jje.go.kr'
}


def parse_school_list_from_excel():
    wb = load_workbook('schoolcodes.xlsx')
    ws = wb['codes']

    SchoolModel.objects.delete()
    # 파싱 전 제거

    for row in range(2, 3587):
        # 현재 학교 코드 엑셀에 있는 row 수
        code = ws['A' + str(row)].value
        # 학교 코드

        region = ws['B' + str(row)].value
        # 교육청

        web_url = WEB_URLS[region]
        # 나이스 URL

        name = ws['C' + str(row)].value
        # 학교 이름

        school = SchoolModel(school_id=code, web_url=web_url, name=name).save()
        ScheduledAlbumModel(school=school).save()
        FreeAlbumModel(school=school).save()

    print('School data Parse Success')


def parse_school_schedules(school_id):
    # school = SchoolModel.objects(school_id=school_id).first()
    #
    # for year in range(2017, 2019):
    #     for month in range(1, 13):
    #         resp = get(_url.format(school.web_url, school.school_id, year, month))
    #         soup = BeautifulSoup(resp.text, 'html.parser')
    #
    #         print(resp.url)
    temp_url = 'http://dsmhs.djsch.kr/scheduleH/list.do?section={}&schdYear={}'
    # TODO 나이스 자체 문제로 인해 학사일정 파싱이 불가능하므로 대마고 기준 데이터로 채움

    schedules = {}

    for year in (2017, 2018):
        for section in (1, 2):
            resp = get(temp_url.format(section, year))
            soup = BeautifulSoup(resp.text, 'html.parser')

            calendars = soup.select('div.calendar')
            for calendar in calendars:
                year, month = re.findall('\d+', calendar.select_one('div span').text)
                schedule_list = calendar.select('li')

                for schedule in schedule_list:
                    day, schedule_name = schedule.text.split('일\xa0:\xa0')

                    if '~' in day:
                        continue

                    schedules['{0}-{1:0>2}-{2:0>2}'.format(year, month, day)] = schedule_name

    school = SchoolModel.objects(school_id='G100000170').first()
    school.update(schedules=schedules)
